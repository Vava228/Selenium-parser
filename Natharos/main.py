import time
from bs4 import BeautifulSoup
import datetime
from selenium import webdriver
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws[f'A1'] = "Дата"
ws[f'B1'] = "Ссылка"
ws[f'C1'] = "Домен"
ws[f'D1'] = "Заголовок"
ws[f'E1'] = "Тематика"
ws[f'F1'] = "Тек.цена"
ws[f'G1'] = "Опт.цена"
ws[f'H1'] = "Блиц"
ws[f'I1'] = "CMS"
ws[f'J1'] = "Траф/сутки"
ws[f'K1'] = "ИКС"
ws[f'L1'] = "Индекс Я"
ws[f'M1'] = "Индекс Г"
ws[f'N1'] = "Доход/мес"

date_today = datetime.datetime.now().strftime('%Y.%m.%d')

headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36'
    }

def get_data():
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(executable_path=r'C:\Freelance\Natharos\chromedriver.exe', options=options)

    try:
        pages_count = int(input('Сколько страниц ты хочешь записать:_'))
    except:
        print('Я думаю ты не понял... Ладно, допустим, 11:')
        pages_count = 10

    hrefs = []
    driver.get(url='https://www.telderi.ru/ru/search/index#page=0&user_id=&website_type%5B0%5D=website&website_type%5B4%5D=website_ib&website_type%5B1%5D=domain&website_type%5B3%5D=youtube&website_type%5B5%5D=vkgroup&website_type%5B6%5D=zenchannel&website_type%5B7%5D=tgchannel&website_type%5B8%5D=instagram&website_type%5B9%5D=mobileapp&website_type%5B10%5D=fbaccount&website_type%5B11%5D=fbgroup&website_type%5B12%5D=fbpage&price%5Bfrom%5D=&price%5Bto%5D=&price_optimal%5Bfrom%5D=&price_optimal%5Bto%5D=&price_bliz%5Bfrom%5D=&price_bliz%5Bto%5D=&revenue%5Btotal%5D%5Bfrom%5D=&revenue%5Btotal%5D%5Bto%5D=&revenue%5Bthousand%5D%5Bfrom%5D=&revenue%5Bthousand%5D%5Bto%5D=&revenue%5Bprofit%5D%5Bfrom%5D=&revenue%5Bprofit%5D%5Bto%5D=&rec%5Bfrom%5D=&rec%5Bto%5D=&seo%5Bview_host%5D%5Bfrom%5D=&seo%5Bview_host%5D%5Bto%5D=&seo%5Bview_hit%5D%5Bfrom%5D=&seo%5Bview_hit%5D%5Bto%5D=&seo%5Bsearch_yandex%5D%5Bfrom%5D=&seo%5Bsearch_yandex%5D%5Bto%5D=&seo%5Bsearch_google%5D%5Bfrom%5D=&seo%5Bsearch_google%5D%5Bto%5D=&seo%5Bindex_ya%5D%5Bfrom%5D=&seo%5Bindex_ya%5D%5Bto%5D=&seo%5Bindex_google%5D%5Bfrom%5D=&seo%5Bindex_google%5D%5Bto%5D=&seo%5Btic%5D%5Bfrom%5D=&seo%5Btic%5D%5Bto%5D=&seo%5Bsqi%5D%5Bfrom%5D=&seo%5Bsqi%5D%5Bto%5D=&seo%5Bpr%5D%5Bfrom%5D=&seo%5Bpr%5D%5Bto%5D=&seo%5Balexa%5D%5Bfrom%5D=&seo%5Balexa%5D%5Bto%5D=&seo%5Blinks%5D%5Bfrom%5D=&seo%5Blinks%5D%5Bto%5D=&soc%5Bsubscribers%5D%5Bfrom%5D=&soc%5Bsubscribers%5D%5Bto%5D=&soc%5Bpost_count%5D%5Bfrom%5D=&soc%5Bpost_count%5D%5Bto%5D=&soc%5Bstat_upload_by%5D=&zen%5Baudience%5D%5Bfrom%5D=&zen%5Baudience%5D%5Bto%5D=&mobileapp%5Binstalls%5D%5Bfrom%5D=&mobileapp%5Binstalls%5D%5Bto%5D=&mobileapp%5Brating%5D%5Bfrom%5D=&mobileapp%5Brating%5D%5Bto%5D=&content=0&age%5Bdomain%5D%5Bfrom%5D=&age%5Bdomain%5D%5Bto%5D=&age%5Bwebsite%5D%5Bfrom%5D=&age%5Bwebsite%5D%5Bto%5D=&domain_type_level=3&reg_panel=&sbk%5Btitle%5D=&sbk%5Bstop_keys_in_title%5D=&sbk%5Bstop_keys_in_desc%5D=&sbk%5Bstop_keys%5D=&sbk%5Burl%5D=&sbk%5Blength%5D%5Bfrom%5D=&sbk%5Blength%5D%5Bto%5D=&sbk%5Bexclude%5D=&started=0&sortby=srart&sorttype=asc')

    for page in range(pages_count):
        try:
            time.sleep(5)
            src = driver.page_source
            driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[3]/div/div/div[2]/div/div/table/tbody/tr[3]/td[2]/div/ul/li[12]/a").click()
        except Exception as ex:
            print(ex)

        soup = BeautifulSoup(src, 'lxml')
        table = soup.find('table', id="sites_list", class_='search_table')
        tags_with_hrefs = table.find_all('tr', class_='website_search_wrap')

        for tag in tags_with_hrefs:
            formatted_tag = tag.find('td', class_='auction_table_title').find('a').get('href')
            hrefs.append(formatted_tag)

    hrefs = list(set(hrefs))
    n = int(2)
    for href in hrefs:
        # Нужно подключить Selenium здесь:
        try:
            driver.get(url=href)
            time.sleep(3)
            src = driver.page_source

        except Exception as ex:
            print(ex)


        soup = BeautifulSoup(src, 'lxml')
        work_zone = soup.find('div', id='auction_left')
        try:
            title = work_zone.find('div', class_='TitleBar_title__2TmIb').find('span').text
        except:
            title = 'Нет!'
        try:
            domain = work_zone.find('div', id='rcmp6').find('span', id='domen_url').find('a', class_='auc_url').get('href')
        except:
            domain = 'Нет доменного имени!'
        try:
            prices = work_zone.find('div', id='rcmp7')
        except Exception as ex:
            print(ex)
        try:
            price_now = prices.find('div', class_='BidBlock_nowPriceBigValue__17j6l').text
        except:
            price_now = 'Нет тек. цены!'
        try:
            required_price = prices.find('div', class_='BidBlock_value__uwFCv').text
        except:
            required_price = 'Нет опт. цены!'
        try:
            blitz_price = prices.find_all('div', class_='BidBlock_value__uwFCv')[1].text
        except:
            blitz_price = 'Нет блиц цены!'
        try:
            e_x = work_zone.find('table', class_='styles_table__33dIu').find_all('tr')[1].find_all('td')[1].find('div').text
        except:
            e_x = 'Нет ИКС!'
        try:
            cms = work_zone.find_all('table', class_='styles_table__33dIu')[1].find_all('tr')[5].find_all('td')[1].text
            if cms == "Да":
                cms = work_zone.find_all('table', class_='styles_table__33dIu')[1].find_all('tr')[6].find_all('td')[1].text
        except:
            cms = 'Нет CMS системы!'
        try:
            profit =  work_zone.find('div', id='rcmp17').find('div', class_='CommonBlock_info__1yu8f CommonBlock_infoWide__1QBVY').find_all('div')[1].find_all('div')[1].text
        except:
            profit = 'Доход не известен!'
        try:
            traffic = work_zone.find('div', id="rcmp14").find('div', class_='CommonBlock_row__3vi2W').find_all('div')[1].text
        except:
            traffic = 'Нет трафика!'
        try:
            theme = work_zone.find('div', id="rcmp11").find_all('table', class_='styles_table__33dIu')[1].find_all('tr')[-1].find_all('td')[1].text
        except:
            theme = 'Нет тематики!'
        try:
            indy = work_zone.find('div', id="rcmp11").find('table', class_='styles_table__33dIu').find_all('tr')[3].find_all('td')[1].find('a').text
        except:
            try:
                indy = work_zone.find('div', id="rcmp11").find('table', class_='styles_table__33dIu').find_all('tr')[3].find_all('td')[1].find('div').text
            except:
                indy = 'Нет индекса Я!'
        try:
            indg = work_zone.find('div', id="rcmp11").find('table', class_='styles_table__33dIu').find_all('tr')[4].find_all('td')[1].find('a').text
        except:
            try:
                indg = work_zone.find('div', id="rcmp11").find('table', class_='styles_table__33dIu').find_all('tr')[4].find_all('td')[1].find('div').text
            except:
                indg = 'Нет индекса Г!'

        ws[f'A{n}'] = date_today
        ws[f'B{n}'] = href
        ws[f'C{n}'] = domain
        ws[f'D{n}'] = title
        ws[f'E{n}'] = theme
        ws[f'F{n}'] = price_now
        ws[f'G{n}'] = required_price
        ws[f'H{n}'] = blitz_price
        ws[f'I{n}'] = cms
        ws[f'J{n}'] = traffic
        ws[f'K{n}'] = e_x
        ws[f'L{n}'] = indy
        ws[f'M{n}'] = indg
        ws[f'N{n}'] = profit
        print(f'Сайт №{n - 1} скопирован!')
        n += 1
    driver.close()
    driver.quit()
def main():
    get_data()
    wb.save(r'data\data.xlsx')
    print('***\nГотово!\n***')

if __name__ == '__main__':
    main()
